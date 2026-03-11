"""
This module provides a number of utility functions.

Functions:
    add_pbs: Read in the PBs and update the Swimmer.
    add_pbs_from_date: Add text about the date PBs were obtained from, to the work sheet.
    add_pbs_last_updated_date: Add the date that the members times file was last modified,
        to the work sheet.
    get_selected_swimmers: Get a list of swimmers.
    get_swimmers: Get all of the swimmers.
"""

from datetime import date, time
from datetime import datetime
import os
import pathlib

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from swimming_stats.constants import (
    Gender,
    LEFT_ALIGNED,
    RIGHT_ALIGNED,
)
from swimming_stats.data import Swimmer
from swimming_stats.settings import (
    ALWAYS_INCLUDE,
    MEMBERS_TIMES_FILE,
    TT_FILE,
    MEMBERS_FILE,
    SWIM_MEMBERSHIPS,
    SQUAD_NAMES,
)
from swimming_stats.swim_manager_settings import *

DATE_FORMAT = "%d/%m/%Y"


def add_pbs_from_date(ws, pb_date, row, title_column, value_column):
    """
    Add text about the date PBs were obtained from, to the work sheet.
    """
    title_cell = ws[f"{title_column}{row}"]
    title_cell.value = "PBs obtained since:"
    title_cell.alignment = RIGHT_ALIGNED
    value_cell = ws[f"{value_column}{row}"]
    value_cell.value = pb_date.strftime(DATE_FORMAT)
    value_cell.alignment = LEFT_ALIGNED


def add_pbs_last_updated_date(ws, row, title_column, value_column):
    """
    Add the date that the members times file was last modified, to the work sheet.
    """
    title_cell = ws[f"{title_column}{row}"]
    title_cell.value = "PBs obtained up to:"
    title_cell.alignment = RIGHT_ALIGNED
    value_cell = ws[f"{value_column}{row}"]
    last_modified = date.fromtimestamp(pathlib.Path(MEMBERS_TIMES_FILE).stat().st_mtime)
    value_cell.value = last_modified.strftime(DATE_FORMAT)
    value_cell.alignment = LEFT_ALIGNED


def get_selected_swimmers(all_swimmers, ignore_squads, ignore_members):
    """
    Get a dict of swimmers.

    @param all_swimmers lst(Swimmer) all the swimmers
    @param ignore_squads lst(str) a list of squads to ignore
    @param ignore_members lst(str) a list of members to ignore

    @return a dict of swimmers, key = full_name, value = Swimmer
    """
    _ignore_squads = []
    for ignore_squad in ignore_squads:
        _ignore_squads.append(SQUAD_NAMES[ignore_squad])

    selected_swimmers = {}
    for swimmer in all_swimmers:
        if swimmer.full_name in ALWAYS_INCLUDE:
            selected_swimmers[swimmer.full_name] = swimmer
            continue
        if swimmer.full_name in ignore_members or swimmer.squad in _ignore_squads:
            continue
        selected_swimmers[swimmer.full_name] = swimmer

    return selected_swimmers


def add_pbs(swimmers, pb_date=None, use_tt=True, only_short_course=False):
    """
    Read in the PBs and update the Swimmer.
    """
    if pb_date is not None:
        print(f"Getting PBs obtained from {pb_date} onwards")

    if use_tt:
        print("Results include time trials")

    print("\nReading data from:")
    print(MEMBERS_TIMES_FILE)
    if use_tt and os.path.isfile(TT_FILE):
        print(TT_FILE)
    print()

    w_book = load_workbook(filename=MEMBERS_TIMES_FILE)
    w_sheet = w_book["Worksheet"]
    _read_pbs_sheet(w_sheet, False, pb_date, only_short_course, swimmers)

    if use_tt and os.path.isfile(TT_FILE):
        w_book = load_workbook(filename=TT_FILE)
        w_sheet = w_book["Worksheet"]
        _read_pbs_sheet(w_sheet, True, pb_date, only_short_course, swimmers)


def get_swimmers(age_at_date=None):
    """
    Get all of the swimmers.
    """
    print("\nReading data from:")
    print(MEMBERS_FILE)
    print()

    excel_wb = MEMBERS_FILE
    w_book = load_workbook(filename=excel_wb)
    w_sheet = w_book["Worksheet"]
    swimmers = _get_swimmers(w_sheet, age_at_date)

    return swimmers


def _read_pbs_sheet(w_sheet, tt, pb_date, only_short_course, swimmers):
    for row in w_sheet.iter_rows(min_row=2):
        if row[0].value is None and row[1].value is None:
            # stop when we get a blank line
            return
        name = f"{row[0].value} {row[1].value}"

        swimmer = swimmers.get(name)
        if swimmer is None:
            continue

        if tt:
            event = row[2].value
            event_time = row[3].value

            if pb_date is not None:
                if isinstance(row[4].value, datetime):
                    event_date = row[4].value
                else:
                    event_date = datetime.strptime(row[4].value, DATE_FORMAT)
                if event_date < pb_date:
                    continue
        else:
            if only_short_course and row[8].value == "LC":
                continue

            if (
                pb_date is not None
                and datetime.strptime(row[5].value, DATE_FORMAT) < pb_date
            ):
                continue

            event = row[6].value
            event_time = row[10].value

        event_time = _convert_event_time(event_time)
        swimmer.add_pb(event, event_time)


def _convert_event_time(event_time):
    if isinstance(event_time, float):
        minutes = 0
        seconds = int(str(event_time).split(".", maxsplit=1)[0])
        decimal = int(str(event_time).split(".")[1])

    elif isinstance(event_time, int):
        minutes = 0
        seconds = event_time
        decimal = 0

    elif isinstance(event_time, str):
        # get string into standard format
        if ":" not in str(event_time):
            minutes = 0
            seconds = int(event_time.split(".")[0])
            decimal = int(event_time.split(".")[1])
        else:
            minutes = int(event_time.split(":")[0])
            seconds = int(event_time.split(":")[1].split(".")[0])
            decimal = int(event_time.split(".")[1])

    if not isinstance(event_time, time) and event_time is not None:
        if len(str(event_time).split(".")) == 1:
            microsecond = 0
        elif len(str(event_time).split(".")[1]) == 1:
            #  i.e. decimal was .1
            microsecond = decimal * 100000
        else:
            # i.e. decimal was .34
            microsecond = decimal * 10000

        event_time = time(minute=minutes, second=seconds, microsecond=microsecond)

    return event_time


def _get_swimmers(w_sheet, age_at_date):
    swimmers = []
    for row in w_sheet.iter_rows(min_row=2):
        if row[MEMBERS_COLUMN_MEMBERSHIP_CATEGORY].value not in SWIM_MEMBERSHIPS:
            continue

        first_name = row[MEMBERS_COLUMN_FIRST_NAME].value
        last_name = row[MEMBERS_COLUMN_LAST_NAME].value
        age = None
        dob = row[MEMBERS_COLUMN_DOB].value
        memb_no = row[MEMBERS_COLUMN_MEMBERSHIP_NUMBER].value
        squad_name = SQUAD_NAMES[row[MEMBERS_COLUMN_SQUAD_NAME].value]

        if age_at_date is not None:
            age = relativedelta(age_at_date, row[MEMBERS_COLUMN_DOB].value.date()).years
        else:
            age = None

        if row[MEMBERS_COLUMN_GENDER].value is None:
            print(f"WARNING: gender not set for {first_name} {last_name}")
            gender = ""
        elif row[MEMBERS_COLUMN_GENDER].value.lower() == "female":
            gender = Gender.FEMALE
        else:
            gender = Gender.MALE

        swimmers.append(
            Swimmer(first_name, last_name, age, dob, gender, memb_no, squad_name)
        )

    return swimmers
