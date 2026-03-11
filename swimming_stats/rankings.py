"""
This module provides the code to Generate rankings based on 200m FC and 200m IM..
"""

from argparse import ArgumentParser, RawDescriptionHelpFormatter
from datetime import datetime, time
from string import ascii_uppercase
import sys

from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle

from swimming_stats.constants import (
    THIN_BORDER,
    BOLD,
    CENTRE_ALIGNED,
    THICK_LEFT_BORDER,
    THICK_RIGHT_BORDER,
    THICK_TOP_BORDER,
    LEFT_ALIGNED,
    RIGHT_ALIGNED,
    RANKING_EVENTS,
    EventType,
)
from swimming_stats.fills import (
    FILL_1,
    FILL_2,
    FILL_3,
    FILL_4,
    FILL_5,
    FILL_6,
    FILL_7,
    FILL_8,
    FILL_9,
    FILL_10,
    FILL_11,
    FILL_12,
    FILL_13,
    FILL_GREEN,
    FILL_RED,
    FILL_EMPTY,
)
from swimming_stats.settings import (
    RANKING_IGNORE_SQUADS,
    RANKING_IGNORE_MEMBERS,
    SQUAD_NAMES,
    SQUAD_FILLS,
    FILE_OUT,
    RANKING_SQUAD_4,
    RANKING_SQUAD_3,
    RANKING_SQUAD_2,
    RANKING_SQUAD_1,
)
from swimming_stats.squad_criteria import squad_criteria
from swimming_stats.utils import (
    add_pbs_from_date,
    add_pbs_last_updated_date,
    get_selected_swimmers,
    get_swimmers,
    add_pbs,
)


# Style for duplicates
DUPLICATES_STYLE = DifferentialStyle(
    font=Font(color="9C0006"), fill=PatternFill(bgColor="FFC7CE")
)

# spreadsheet columns
RANK_COLUMN = "A"
FIRST_NAME_COLUMN = "B"
LAST_NAME_COLUMN = "C"
SQUAD_COLUMN = "D"
RESULTS_START_COLUMN = "I"
RESULTS_START_COLUMN_NO = 9
RESULTS_END_COLUMN_1 = "Q"
RESULTS_END_COLUMN_2 = "S"
RESULTS_END_COLUMN = "Z"
CRITERIA_START_COLUMN = "AA"
LAST_COLUMN = "BP"
DATE_TITLE_COLUMN = "X"
DATE_COLUMNS = "YZ"
SQUAD_SCORES_COLUMNS = "EFGH"
RESULTS = "IJKLMNOPQRS"
HEADER_ROW = "1"


def _add_rankings(swimmers, event):
    """
    Add a ranking to the members in results for the given event.

    The swimmers object is updated.

    @param swimmers
    @param event

    """
    # a dict of swimmers and their time for the event
    swimmers_with_time = {}

    # a list of swimmers with no times for this event
    no_result = []

    for swimmer in swimmers.values():
        if swimmer.pbs.get(event) is None:
            #  the swimmer does not have a time for this event
            no_result.append(swimmer.full_name)
        else:
            swimmers_with_time[swimmer.full_name] = swimmer.pbs.get(event)

    ranked = sorted(swimmers_with_time.items(), key=_value_getter)
    last_value = None
    last_rank = 1

    for i, (member, value) in enumerate(ranked):
        if value == last_value:
            rank = last_rank
        else:
            rank = i + 1
        last_rank = rank
        last_value = value

        swimmers[member].add_ranking(event, rank)

    total_members = len(swimmers)
    for member in no_result:
        swimmers[member].add_ranking(event, total_members)


def _add_overall_ranking(swimmers, events):
    """
    Add a ranking to each member based on the given events.

    @param swimmers
    @param events list(str) the event used for ranking

    """
    rankings = {}
    for name, swimmer in swimmers.items():
        ranking = 0
        for event in events:
            ranking = ranking + swimmer.rankings[event]
        rankings[name] = ranking

    ranked = sorted(rankings.items(), key=_value_getter)
    last_value = None
    last_rank = 1

    for i, (swimmer, value) in enumerate(ranked):
        if value == last_value:
            rank = last_rank
        else:
            rank = i + 1
        last_rank = rank
        last_value = value

        swimmers[swimmer].add_ranking("ranking", rank)


def _value_getter(item):
    return item[1]


def _ranking_getter(item):
    return item[1].rankings["ranking"]


def _write_squad_criteria(wb):
    """ "
    Write the squad criteria sheet.

    """
    ws = wb.create_sheet("Squad Criteria")

    ws.cell(row=1, column=1).value = "Squad"
    ws.cell(row=1, column=2).value = "50m FC"
    ws.cell(row=1, column=3).value = "200m FC"
    ws.cell(row=1, column=4).value = "50m BK"
    ws.cell(row=1, column=5).value = "200m BK"
    ws.cell(row=1, column=6).value = "50m Br"
    ws.cell(row=1, column=7).value = "200m BR"
    ws.cell(row=1, column=8).value = "50m Fly"
    ws.cell(row=1, column=9).value = "200m Fly"
    ws.cell(row=1, column=10).value = "200m IM"
    ws.cell(row=1, column=11).value = "400 FC"
    ws.cell(row=1, column=12).value = "400 IM"

    for row, squad in enumerate(squad_criteria):
        row = row + 2
        ws.cell(row=row, column=1).value = squad

        for column, time_value in enumerate(squad_criteria[squad]):
            ws.cell(row=row, column=column + 2).value = time(
                minute=time_value[0], second=time_value[1]
            )

    rows = ws["A1:L1"]
    for row in rows:
        for cell in row:
            cell.border = THIN_BORDER

    # Add some colour
    for row, fill in enumerate([FILL_2, FILL_3, FILL_4, FILL_6, FILL_7]):
        rows = ws[f"A{row+2}:L{row+2}"]
        for row in rows:
            for cell in row:
                cell.fill = fill
                cell.border = THIN_BORDER

    # Format numbers
    rows = ws["B2:L6"]
    for row in rows:
        for cell in row:
            cell.number_format = "mm:ss"
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_squads(wb, swimmers):
    """
    Write the squad lists to a sheet.
    """
    ws = wb.create_sheet("Squads")

    squads = {}
    for swimmer in swimmers:
        if squads.get(swimmer.squad) is None:
            squads[swimmer.squad] = [swimmer.full_name]
        else:
            squads[swimmer.squad].append(swimmer.full_name)

    letters = dict(enumerate(ascii_uppercase, start=1))

    for i, (squad, fill) in enumerate(SQUAD_FILLS.items()):
        column_width = len(squad)
        ws.cell(row=1, column=i + 1).value = squad
        ws.cell(row=1, column=i + 1).border = THIN_BORDER
        ws.cell(row=1, column=i + 1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=1, column=i + 1).fill = fill

        if squads.get(SQUAD_NAMES[squad]) is None:
            continue
        for j, name in enumerate(squads.get(SQUAD_NAMES[squad])):
            column_width = max(column_width, len(name))
            ws.cell(row=j + 2, column=i + 1).value = name
            ws.cell(row=j + 2, column=i + 1).alignment = Alignment(
                horizontal="left", vertical="center"
            )
            ws.cell(row=j + 2, column=i + 1).border = THIN_BORDER
            ws.cell(row=j + 2, column=i + 1).fill = fill

        ws.column_dimensions[letters[i + 1]].width = column_width


def _write_rankings(ws, swimmers, pb_date, tt_included):
    """
    Write the rankings sheet.

    @param ws work sheet
    @param swimmers
    @param pb_date
    @param tt_included(bool) time trial results have been included

    """

    # Header row
    headers = [
        "Rank",
        "First Name",
        "Last Name",
        "Squad",
        RANKING_SQUAD_4,
        RANKING_SQUAD_3,
        RANKING_SQUAD_2,
        RANKING_SQUAD_1,
    ]

    for event in RANKING_EVENTS:
        try:
            headers.append(event.value)
        except AttributeError:
            headers.append(event)

    for i, value in enumerate(headers):
        ws.cell(row=1, column=1 + i).value = value
        ws.cell(row=1, column=1 + i).font = BOLD

    # Results
    for i, swimmer in enumerate(swimmers.values()):
        ws.cell(row=i + 2, column=1).value = swimmer.rankings["ranking"]
        ws.cell(row=i + 2, column=2).value = swimmer.first_name
        ws.cell(row=i + 2, column=3).value = swimmer.last_name
        ws.cell(row=i + 2, column=4).value = swimmer.squad

        for j, event in enumerate(RANKING_EVENTS):
            if swimmer.pbs.get(event) is not None:
                ws.cell(row=i + 2, column=j + RESULTS_START_COLUMN_NO).value = (
                    swimmer.pbs.get(event)
                )

        _add_squad_scores(ws, i + 2)

    last_row = i + 2

    # Highlight duplicates
    rule = Rule(type="duplicateValues", text="highlight", dxf=DUPLICATES_STYLE)
    ws.conditional_formatting.add(f"{RANK_COLUMN}2:{RANK_COLUMN}{last_row}", rule)

    _highlight_squad_scores(ws, last_row)

    _highlight_squads(ws, last_row)

    _highlight_times(ws, last_row)

    _add_borders_and_alignment(ws, last_row)

    # Format numbers
    rows = ws[f"{RESULTS_START_COLUMN}{HEADER_ROW}:{RESULTS_END_COLUMN}{last_row}"]
    for row in rows:
        for cell in row:
            cell.number_format = "mm:ss.00"

    # Column widths for name columns
    ws.column_dimensions[FIRST_NAME_COLUMN].width = 15
    ws.column_dimensions[LAST_NAME_COLUMN].width = 15
    ws.column_dimensions[SQUAD_COLUMN].width = 10
    for column in "EFGH":
        ws.column_dimensions[column].width = 5
    for column in "AIJKLMNOPQRSTUVWXYZ":
        ws.column_dimensions[column].width = 10

    # Add  sorting
    ws.auto_filter.ref = f"{RANK_COLUMN}{HEADER_ROW}:{LAST_COLUMN}{last_row}"
    ws.auto_filter.add_sort_condition(
        f"{RANK_COLUMN}{HEADER_ROW}:{RANK_COLUMN}{last_row}"
    )

    i += 4

    _add_squad_key(ws, i)
    _add_squad_scores_key(ws, i)

    if pb_date is not None:
        add_pbs_from_date(ws, pb_date, i, "B", "C")
        i += 1

    add_pbs_last_updated_date(ws, i, "B", "C")
    i += 1

    if tt_included:
        ws.cell(row=i, column=2).value = "Times from time trials have been included"
        i += 1

    i += 1
    _add_ranking_text(ws, i)

    # Hide
    ws.column_dimensions.group(
        start=CRITERIA_START_COLUMN, end=LAST_COLUMN, hidden=True
    )


def _add_borders_and_alignment(ws, last_row):
    """
    Add borders and alignment to the table.

    @param ws(): the work sheet
    @param last_row(int): the last row of the table

    """
    rows = ws[f"{RANK_COLUMN}{HEADER_ROW}:{RESULTS_END_COLUMN}{last_row}"]
    for row in rows:
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = CENTRE_ALIGNED

        for column in [0, 1, 4, 8, 19]:
            row[column].border = THICK_LEFT_BORDER
        row[25].border = THICK_RIGHT_BORDER

    row = ws[f"{RANK_COLUMN}{last_row+1}:{RESULTS_END_COLUMN}{last_row+1}"][0]
    for cell in row:
        cell.border = THICK_TOP_BORDER

    rows = ws[f"{SQUAD_COLUMN}{HEADER_ROW}:{RESULTS_END_COLUMN}{HEADER_ROW}"]
    for row in rows:
        for cell in row:
            cell.alignment = LEFT_ALIGNED


def _add_ranking_text(ws, row):
    """
    Add text about the rankings on thhe  given row.

    @param ws worksheet
    @param row worksheet row

    """
    ws.cell(row=row, column=2).value = "Ranking is based on:"
    ws.cell(row=row, column=2).alignment = RIGHT_ALIGNED
    ws.cell(row=row, column=3).value = "200m FC"
    ws.cell(row=row + 1, column=3).value = "200m IM"


def _add_squad_scores(ws, row):
    letters = dict(enumerate(ascii_uppercase, start=0))

    # RANKING_SQUAD_4 scores
    next_column = len(RANKING_EVENTS) + RESULTS_START_COLUMN_NO
    for column_no in range(1, 10):
        ws.cell(row=row, column=next_column + column_no).value = (
            f"=IF(AND({RESULTS[column_no-1]}{row}<='Squad Criteria'!{letters[column_no]}$5,"
            + f'{RESULTS[column_no-1]}{row}<>""),1,0)'
        )
    ws.cell(row=row, column=5).value = f"=sum(AB{row}:AJ{row})"

    # RANKING_SQUAD_3 scores
    next_column = next_column + 10
    for column_no in range(1, 10):
        ws.cell(row=row, column=next_column + column_no).value = (
            f"=IF(AND({RESULTS[column_no-1]}{row}<='Squad Criteria'!{letters[column_no]}$4,"
            + f'{RESULTS[column_no-1]}{row}<>""),1,0)'
        )
    ws.cell(row=row, column=6).value = f"=sum(AL{row}:AT{row})"

    # RANKING_SQUAD_2 scores
    next_column = next_column + 10
    for column_no in range(1, 10):
        ws.cell(row=row, column=next_column + column_no).value = (
            f"=IF(AND({RESULTS[column_no-1]}{row}<='Squad Criteria'!{letters[column_no]}$3,"
            + f'{RESULTS[column_no-1]}{row}<>""),1,0)'
        )
    ws.cell(row=row, column=7).value = f"=sum(AV{row}:BD{row})"

    # RANKING_SQUAD_1 scores
    next_column = next_column + 10
    for column_no in range(1, 12):
        ws.cell(row=row, column=next_column + column_no).value = (
            f"=IF(AND({RESULTS[column_no-1]}{row}<='Squad Criteria'!{letters[column_no]}$2,"
            + f'{RESULTS[column_no-1]}{row}<>""),1,0)'
        )
    ws.cell(row=row, column=8).value = f"=sum(BF{row}:{LAST_COLUMN}{row})"


def _add_squad_key(ws, start_row):
    # Add a key
    ws.merge_cells(f"J{start_row}:O{start_row}")
    merged = ws[f"J{start_row}"]
    merged.value = "Squads and squad times"
    merged.border = THIN_BORDER
    merged.alignment = CENTRE_ALIGNED
    for column in "JKLMNO":
        cell = ws[f"{column}{start_row}"]
        cell.border = THIN_BORDER

    for j, (squad, fill) in enumerate(SQUAD_FILLS.items()):
        if j > 9:
            break
        key_row = start_row + j + 1
        if j < 5:
            key_columns = "JKL"
        else:
            key_row = key_row - 5
            key_columns = "MNO"
        ws.merge_cells(f"{key_columns[0]}{key_row}:{key_columns[-1]}{key_row}")
        merged = ws[f"{key_columns[0]}{key_row}"]
        merged.value = squad
        merged.fill = fill
        merged.border = THIN_BORDER
        merged.alignment = CENTRE_ALIGNED

        # shouldn't have to apply border to each cell
        for column in key_columns:
            cell = ws[f"{column}{key_row}"]
            cell.border = THIN_BORDER


def _add_squad_scores_key(ws, start_row):
    # Add a key
    ws.merge_cells(
        f"{SQUAD_SCORES_COLUMNS[0]}{start_row}:{SQUAD_SCORES_COLUMNS[-1]}{start_row}"
    )
    merged = ws[f"{SQUAD_SCORES_COLUMNS[0]}{start_row}"]
    merged.value = "Meets squad criteria"
    merged.alignment = CENTRE_ALIGNED

    for column in SQUAD_SCORES_COLUMNS:
        cell = ws[f"{column}{start_row}"]
        cell.border = THIN_BORDER

    row = start_row + 1
    ws.merge_cells(f"{SQUAD_SCORES_COLUMNS[0]}{row}:{SQUAD_SCORES_COLUMNS[-1]}{row}")
    merged = ws[f"{SQUAD_SCORES_COLUMNS[0]}{row}"]
    merged.value = "YES"
    merged.alignment = CENTRE_ALIGNED
    merged.fill = FILL_GREEN

    for column in SQUAD_SCORES_COLUMNS:
        cell = ws[f"{column}{row}"]
        cell.border = THIN_BORDER

    row = row + 1
    ws.merge_cells(f"{SQUAD_SCORES_COLUMNS[0]}{row}:{SQUAD_SCORES_COLUMNS[-1]}{row}")
    merged = ws[f"{SQUAD_SCORES_COLUMNS[0]}{row}"]
    merged.value = "NO"
    merged.alignment = CENTRE_ALIGNED
    merged.fill = FILL_RED

    for column in SQUAD_SCORES_COLUMNS:
        cell = ws[f"{column}{row}"]
        cell.border = THIN_BORDER


def _highlight_squads(ws, last_row):
    # Highlight squads
    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$A$2:$A$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_1,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$B$2:$B$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_2,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$C$2:$C$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_3,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$D$2:$D$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_4,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$E$2:$E$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_5,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$F$2:$F$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_6,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$G$2:$G$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_7,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$H$2:$H$20,0)"
        ],
        stopIfTrue=True,
        fill=FILL_8,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$I$2:$I$20, 0)"
        ],
        stopIfTrue=True,
        fill=FILL_9,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$J$2:$J$20, 0)"
        ],
        stopIfTrue=True,
        fill=FILL_10,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$K$2:$K$20, 0)"
        ],
        stopIfTrue=True,
        fill=FILL_11,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$L$2:$L$20, 0)"
        ],
        stopIfTrue=True,
        fill=FILL_12,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$M$2:$M$20, 0)"
        ],
        stopIfTrue=True,
        fill=FILL_13,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )

    rule = FormulaRule(
        formula=[
            f'MATCH(CONCATENATE(${FIRST_NAME_COLUMN}2," ",${LAST_NAME_COLUMN}2),'
            + "Squads!$N$2:$N$20, 0)"
        ],
        stopIfTrue=True,
        fill=FILL_EMPTY,
    )
    ws.conditional_formatting.add(
        f"{FIRST_NAME_COLUMN}2:{SQUAD_COLUMN}{last_row}", rule
    )


def _highlight_squad_scores(ws, last_row):
    # Highlight RANKING_SQUAD_4 count
    rule = CellIsRule(
        operator="greaterThan",
        formula=["5"],
        stopIfTrue=True,
        fill=FILL_GREEN,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[0]}2:{SQUAD_SCORES_COLUMNS[0]}{last_row}", rule
    )

    rule = CellIsRule(
        operator="lessThan",
        formula=["6"],
        stopIfTrue=True,
        fill=FILL_RED,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[0]}2:{SQUAD_SCORES_COLUMNS[0]}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_3 count
    rule = CellIsRule(
        operator="greaterThan",
        formula=["5"],
        stopIfTrue=True,
        fill=FILL_GREEN,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[1]}2:{SQUAD_SCORES_COLUMNS[1]}{last_row}", rule
    )

    rule = CellIsRule(
        operator="lessThan",
        formula=["6"],
        stopIfTrue=True,
        fill=FILL_RED,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[1]}2:{SQUAD_SCORES_COLUMNS[1]}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_2 count
    rule = CellIsRule(
        operator="greaterThan",
        formula=["5"],
        stopIfTrue=True,
        fill=FILL_GREEN,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[2]}2:{SQUAD_SCORES_COLUMNS[2]}{last_row}", rule
    )

    rule = CellIsRule(
        operator="lessThan",
        formula=["6"],
        stopIfTrue=True,
        fill=FILL_RED,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[2]}2:{SQUAD_SCORES_COLUMNS[2]}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_1 count
    rule = CellIsRule(
        operator="greaterThan",
        formula=["5"],
        stopIfTrue=True,
        fill=FILL_GREEN,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[3]}2:{SQUAD_SCORES_COLUMNS[3]}{last_row}", rule
    )

    rule = CellIsRule(
        operator="lessThan",
        formula=["6"],
        stopIfTrue=True,
        fill=FILL_RED,
    )
    ws.conditional_formatting.add(
        f"{SQUAD_SCORES_COLUMNS[3]}2:{SQUAD_SCORES_COLUMNS[3]}{last_row}", rule
    )


def _highlight_times(ws, last_row):
    # Highlight if empty
    rule = CellIsRule(
        operator="equal",
        formula=["'Squad Criteria'!B$7"],
        stopIfTrue=True,
        fill=FILL_EMPTY,
    )
    ws.conditional_formatting.add(
        f"{RESULTS_START_COLUMN}2:{RESULTS_END_COLUMN}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_1 times
    rule = CellIsRule(
        operator="lessThanOrEqual",
        formula=["'Squad Criteria'!B$2"],
        stopIfTrue=True,
        fill=FILL_2,
    )
    ws.conditional_formatting.add(
        f"{RESULTS_START_COLUMN}2:{RESULTS_END_COLUMN_2}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_2 times
    rule = CellIsRule(
        operator="lessThanOrEqual",
        formula=["'Squad Criteria'!B$3"],
        stopIfTrue=True,
        fill=FILL_3,
    )
    ws.conditional_formatting.add(
        f"{RESULTS_START_COLUMN}2:{RESULTS_END_COLUMN_1}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_3 times
    rule = CellIsRule(
        operator="lessThanOrEqual",
        formula=["'Squad Criteria'!B$4"],
        stopIfTrue=True,
        fill=FILL_4,
    )
    ws.conditional_formatting.add(
        f"{RESULTS_START_COLUMN}2:{RESULTS_END_COLUMN_1}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_4 times
    rule = CellIsRule(
        operator="lessThanOrEqual",
        formula=["'Squad Criteria'!B$5"],
        stopIfTrue=True,
        fill=FILL_6,
    )
    ws.conditional_formatting.add(
        f"{RESULTS_START_COLUMN}2:{RESULTS_END_COLUMN_1}{last_row}", rule
    )

    # Highlight RANKING_SQUAD_5 times
    rule = CellIsRule(
        operator="lessThanOrEqual",
        formula=["'Squad Criteria'!B$6"],
        stopIfTrue=True,
        fill=FILL_7,
    )
    ws.conditional_formatting.add(
        f"{RESULTS_START_COLUMN}2:{RESULTS_END_COLUMN_1}{last_row}", rule
    )


def _parse_command_line(argv):
    parser = ArgumentParser(
        description="Generate rankings based on 200m FC and 200m IM",
        formatter_class=RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "-p", "--pb_date", help="Only use PBs from this date onwards (dd-mm-yyyy)"
    )
    parser.set_defaults(pb_date=None)

    parser.add_argument(
        "-s",
        "--short_course_only",
        action="store_true",
        help="Only use short course times",
    )
    parser.set_defaults(short_course=False)

    parser.add_argument(
        "-t",
        "--time_trial",
        action="store_true",
        help="Use times from the time trial file",
    )
    parser.set_defaults(time_trial=False)

    return parser.parse_args(argv[1:])


def main():
    """
    Generate the ranking.
    """
    args = _parse_command_line(sys.argv)

    pb_date = args.pb_date
    if pb_date is not None:
        print(f"\nOnly PBs from {pb_date} onwards will be used")
        pb_date = datetime.strptime(pb_date, "%d-%m-%Y")

    short_course_only = args.short_course_only
    if short_course_only:
        print("\nOnly short course results will be included")

    use_tt = args.time_trial
    if use_tt:
        print("\nTime trial results will be included")

    all_swimmers = get_swimmers()
    selected_swimmers = get_selected_swimmers(
        all_swimmers, RANKING_IGNORE_SQUADS, RANKING_IGNORE_MEMBERS
    )

    add_pbs(
        selected_swimmers,
        use_tt=use_tt,
        pb_date=pb_date,
        only_short_course=short_course_only,
    )

    events = [EventType.FC_200, EventType.IM_200]
    # events = [
    #     EventType.BACK_50,
    #     EventType.BACK_100,
    #     EventType.BACK_200,
    #     EventType.BREAST_50,
    #     EventType.BREAST_100,
    #     EventType.BREAST_200,
    #     EventType.FLY_50,
    #     EventType.FLY_100,
    #     EventType.FLY_200,
    #     EventType.FC_50,
    #     EventType.FC_100,
    #     EventType.FC_200,
    #     EventType.FC_400,
    #     EventType.FC_800,
    #     EventType.FC_1500,
    #     EventType.IM_100,
    #     EventType.IM_200,
    #     EventType.IM_400,
    # ]
    for event in events:
        _add_rankings(selected_swimmers, event)

    _add_overall_ranking(selected_swimmers, events)

    ranked_results = dict(sorted(selected_swimmers.items(), key=_ranking_getter))

    wb = Workbook()
    ws = wb.active
    ws.title = "Rankings"
    _write_rankings(ws, ranked_results, pb_date, use_tt)

    _write_squad_criteria(wb)
    _write_squads(wb, all_swimmers)

    wb.save(FILE_OUT)
    print(f"\nWritten:\n{FILE_OUT}\n")


if __name__ == "__main__":
    main()
