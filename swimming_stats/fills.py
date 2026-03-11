"""
This module contains the fills for the squads.
"""

from openpyxl.styles import Color, PatternFill


__all__ = [
    "FILL_1",
    "FILL_2",
    "FILL_3",
    "FILL_4",
    "FILL_5",
    "FILL_6",
    "FILL_7",
    "FILL_8",
    "FILL_9",
    "FILL_10",
    "FILL_11",
    "FILL_12",
    "FILL_13",
    "FILL_EMPTY",
    "FILL_GREEN",
    "FILL_RED",
]

# Colours and fills
# Define the colours to be used for the squads
COLOUR_1 = Color("B96029")
COLOUR_2 = Color("FF3843")
COLOUR_3 = Color("F4B084")
COLOUR_4 = Color("EDBED2")
COLOUR_5 = Color("FFC004")
COLOUR_6 = Color("FFE699")
COLOUR_6 = Color("ffffff")
COLOUR_6 = Color("00B0F0")
COLOUR_7 = Color("8EE9EE")
COLOUR_8 = Color("00AFAF")
COLOUR_9 = Color("C6E0B4")
COLOUR_10 = Color("E2EFDA")
COLOUR_11 = Color("eef4e3")
COLOUR_12 = Color("A6A6A6")
COLOUR_13 = Color("bfbfbf")
COLOUR_EMPTY = Color("ffffff")
COLOUR_GREEN = Color("00ff00")
COLOUR_RED = Color("ff3300")

# Define the colours to be used for Arena League sRGB
COLOUR_ARENA_BLUE = Color("00CCFF")
COLOUR_ARENA_BROWN = Color("FFCC99")
COLOUR_ARENA_GREEN = Color("02FF00")
COLOUR_ARENA_ORANGE = Color("FEC000")

# Squad specific fills
FILL_1 = PatternFill(fill_type="solid", start_color=COLOUR_1, end_color=COLOUR_1)
FILL_2 = PatternFill(fill_type="solid", start_color=COLOUR_2, end_color=COLOUR_2)
FILL_3 = PatternFill(fill_type="solid", start_color=COLOUR_3, end_color=COLOUR_3)
FILL_4 = PatternFill(fill_type="solid", start_color=COLOUR_4, end_color=COLOUR_4)
FILL_5 = PatternFill(fill_type="solid", start_color=COLOUR_5, end_color=COLOUR_5)
FILL_6 = PatternFill(fill_type="solid", start_color=COLOUR_6, end_color=COLOUR_6)
FILL_7 = PatternFill(fill_type="solid", start_color=COLOUR_7, end_color=COLOUR_7)
FILL_8 = PatternFill(fill_type="solid", start_color=COLOUR_8, end_color=COLOUR_8)
FILL_9 = PatternFill(fill_type="solid", start_color=COLOUR_9, end_color=COLOUR_9)
FILL_10 = PatternFill(fill_type="solid", start_color=COLOUR_10, end_color=COLOUR_10)
FILL_11 = PatternFill(fill_type="solid", start_color=COLOUR_11, end_color=COLOUR_11)
FILL_12 = PatternFill(fill_type="solid", start_color=COLOUR_12, end_color=COLOUR_12)
FILL_13 = PatternFill(fill_type="solid", start_color=COLOUR_13, end_color=COLOUR_13)
FILL_EMPTY = PatternFill(
    fill_type="solid", start_color=COLOUR_EMPTY, end_color=COLOUR_EMPTY
)
FILL_GREEN = PatternFill(
    fill_type="solid", start_color=COLOUR_GREEN, end_color=COLOUR_GREEN
)
FILL_RED = PatternFill(fill_type="solid", start_color=COLOUR_RED, end_color=COLOUR_RED)

# Arena League specific fills
FILL_ARENA_BLUE = PatternFill(
    fill_type="solid", start_color=COLOUR_ARENA_BLUE, end_color=COLOUR_ARENA_BLUE
)
FILL_ARENA_BROWN = PatternFill(
    fill_type="solid", start_color=COLOUR_ARENA_BLUE, end_color=COLOUR_ARENA_BROWN
)
FILL_ARENA_GREEN = PatternFill(
    fill_type="solid", start_color=COLOUR_ARENA_GREEN, end_color=COLOUR_ARENA_GREEN
)
FILL_ARENA_ORANGE = PatternFill(
    fill_type="solid", start_color=COLOUR_ARENA_ORANGE, end_color=COLOUR_ARENA_ORANGE
)
